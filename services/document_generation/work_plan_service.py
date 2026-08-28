"""Cronograma deterministico y PDF del plan de trabajo GCDTP-F-021."""

from __future__ import annotations

import math
import platform
import re
import shutil
import tempfile
import warnings
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from config.settings import SENA_LOGO_PATH, WORK_PLAN_TEMPLATE
from services.document_generation.spreadsheet_assets import add_centered_image
from services.document_generation.spreadsheet_pdf import (
    SpreadsheetPdfError,
    convert_xlsx_to_pdf,
)


class WorkPlanError(RuntimeError):
    """Error controlado al calcular o generar el plan de trabajo."""


@dataclass(frozen=True)
class ScheduledActivity:
    code: str
    description: str
    phase: int
    row: int
    category: str
    start_date: date
    end_date: date

    @property
    def duration(self) -> int:
        return (self.end_date - self.start_date).days + 1

    @property
    def date_label(self) -> str:
        start = self.start_date.strftime("%d/%m/%Y")
        if self.end_date == self.start_date:
            return start
        return f"{start} - {self.end_date.strftime('%d/%m/%Y')}"


class WorkPlanService:
    """Calcula, valida y diligencia el formato institucional GCDTP-F-021."""

    SHEET_NAME = "GCDTP-F-021"
    ACTIVITY_ROWS = {
        "1.1": 18,
        "1.2": 19,
        "1.3": 20,
        "1.4": 21,
        "1.5": 22,
        "2.1": 24,
        "2.2": 25,
        "2.3": 26,
        "2.4": 27,
        "2.5": 28,
        "2.6": 29,
        "2.7": 30,
        "2.8": 31,
        "2.9": 32,
        "3.1": 34,
        "3.2": 35,
        "3.3": 36,
        "3.4": 37,
        "3.5": 38,
        "3.6": 39,
        "3.7": 40,
        "3.8": 41,
        "4.1": 43,
        "4.2": 44,
        "4.3": 45,
        "4.4": 46,
        "4.5": 47,
        "4.6": 48,
    }
    PHASE_CODES = {
        1: ("1.1", "1.2", "1.3", "1.4", "1.5"),
        2: ("2.1", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7", "2.8", "2.9"),
        3: ("3.1", "3.2", "3.3", "3.4", "3.5", "3.6", "3.7", "3.8"),
        4: ("4.1", "4.2", "4.3", "4.4", "4.5", "4.6"),
    }
    PHASE_WEIGHTS = (0.05, 0.05, 0.80, 0.10)
    MIN_PHASE_LENGTHS = (2, 3, 4, 2)
    MAX_DAILY_ACTIVITIES = 3
    MAX_ACTIVITY_DURATION = 3
    START_CODES = {"1.1", "2.1", "3.1", "4.1"}
    FINAL_CODES = {"1.4", "2.8", "3.7", "4.5"}
    CLOSING_PAIRS = {
        "1.4": "1.5",
        "2.8": "2.9",
        "3.7": "3.8",
        "4.5": "4.6",
    }

    def build_schedule(self, project: dict[str, Any]) -> list[ScheduledActivity]:
        start_date, end_date, _, _ = self.validate_project(project)
        descriptions = self._load_descriptions()
        total_days = (end_date - start_date).days + 1
        phase_lengths = self._allocate_phase_lengths(total_days)
        schedule: list[ScheduledActivity] = []
        phase_start = start_date

        for phase, phase_length in enumerate(phase_lengths, start=1):
            phase_end = phase_start + timedelta(days=phase_length - 1)
            codes = self.PHASE_CODES[phase]
            closing_code = codes[-2]
            closing_pair = codes[-1]
            schedule.append(
                self._activity(codes[0], descriptions, phase, "Inicio", phase_start, phase_start)
            )
            middle_codes = codes[1:-2]
            middle_scheduler = (
                self._schedule_execution_middle if phase == 3 else self._schedule_middle
            )
            schedule.extend(
                middle_scheduler(
                    middle_codes, descriptions, phase, phase_start, phase_end
                )
            )
            schedule.append(
                self._activity(
                    closing_code,
                    descriptions,
                    phase,
                    "Cierre",
                    phase_end,
                    phase_end,
                )
            )
            schedule.append(
                self._activity(
                    closing_pair,
                    descriptions,
                    phase,
                    "Cierre",
                    phase_end,
                    phase_end,
                )
            )
            phase_start = phase_end + timedelta(days=1)

        schedule.sort(key=lambda item: self.ACTIVITY_ROWS[item.code])
        self.validate_schedule(schedule, start_date, end_date)
        return schedule

    def generate(
        self,
        project: dict[str, Any],
        schedule: list[ScheduledActivity],
    ) -> tuple[bytes, str]:
        start_date, end_date, talent, expert = self.validate_project(project)
        self.validate_schedule(schedule, start_date, end_date)
        with tempfile.TemporaryDirectory(prefix="tp_work_plan_") as temp_directory:
            directory = Path(temp_directory)
            xlsx_path = directory / "plan_trabajo.xlsx"
            pdf_path = directory / "plan_trabajo.pdf"
            shutil.copy2(WORK_PLAN_TEMPLATE, xlsx_path)
            if platform.system() == "Windows":
                self._complete_workbook_with_excel(
                    xlsx_path, project, schedule, talent, expert
                )
            else:
                self._complete_workbook(xlsx_path, project, schedule, talent, expert)
            try:
                convert_xlsx_to_pdf(
                    xlsx_path,
                    pdf_path,
                    sheet_name=self.SHEET_NAME,
                )
            except SpreadsheetPdfError as error:
                raise WorkPlanError(str(error)) from error
            pdf_data = pdf_path.read_bytes()
        code = self._safe_filename(str(project.get("code") or "proyecto"))
        return pdf_data, f"Plan de Trabajo {code}.pdf"

    def validate_project(
        self,
        project: dict[str, Any],
    ) -> tuple[date, date, dict[str, Any], dict[str, Any]]:
        if not WORK_PLAN_TEMPLATE.is_file():
            raise WorkPlanError("No se encontro la plantilla GCDTP-F-021.")
        start_date = self._as_date(project.get("start_date"))
        end_date = self._as_date(project.get("end_date"))
        if start_date is None:
            raise WorkPlanError("El proyecto no tiene fecha de inicio.")
        if end_date is None:
            raise WorkPlanError("El proyecto no tiene fecha de cierre.")
        if end_date < start_date:
            raise WorkPlanError("La fecha de cierre no puede ser anterior a la fecha de inicio.")
        minimum_days = sum(self.MIN_PHASE_LENGTHS)
        if (end_date - start_date).days + 1 < minimum_days:
            raise WorkPlanError(
                f"El proyecto debe disponer de al menos {minimum_days} dias para distribuir "
                "las actividades sin superar tres actividades por dia."
            )
        talents = project.get("talents", [])
        talent = next((item for item in talents if item.get("role") == "ejecutor"), None)
        if talent is None:
            talent = next((item for item in talents if item.get("role") == "titular"), None)
        if talent is None:
            raise WorkPlanError("El proyecto requiere un talento ejecutor o titular.")
        expert = project.get("expert")
        if not expert or not expert.get("name"):
            raise WorkPlanError("El proyecto requiere un experto Tecnoparque.")
        return start_date, end_date, talent, expert

    def validate_schedule(
        self,
        schedule: list[ScheduledActivity],
        project_start: date,
        project_end: date,
    ) -> None:
        by_code = {item.code: item for item in schedule}
        missing = set(self.ACTIVITY_ROWS) - set(by_code)
        if missing:
            raise WorkPlanError("Faltan actividades en el cronograma: " + ", ".join(sorted(missing)))
        for item in schedule:
            if item.start_date < project_start or item.end_date > project_end:
                raise WorkPlanError(f"La actividad {item.code} esta fuera de las fechas del proyecto.")
            if item.end_date < item.start_date:
                raise WorkPlanError(f"La actividad {item.code} tiene fechas invalidas.")
            if item.phase != 3 and item.duration > self.MAX_ACTIVITY_DURATION:
                raise WorkPlanError(
                    f"La actividad {item.code} no puede superar tres dias."
                )
        if by_code["1.1"].start_date != project_start:
            raise WorkPlanError("La actividad 1.1 debe iniciar con el proyecto.")
        if by_code["4.5"].start_date != project_end or by_code["4.6"].start_date != project_end:
            raise WorkPlanError("Las actividades 4.5 y 4.6 deben coincidir con el cierre del proyecto.")
        for code in self.START_CODES | self.FINAL_CODES:
            if by_code[code].duration != 1:
                raise WorkPlanError(f"La actividad {code} debe durar exactamente un dia.")
        for first, second in self.CLOSING_PAIRS.items():
            left, right = by_code[first], by_code[second]
            if left.start_date != right.start_date or left.end_date != right.end_date:
                raise WorkPlanError(f"Las actividades {first} y {second} deben compartir fecha.")
        current_date = project_start
        while current_date <= project_end:
            active = [
                item.code
                for item in schedule
                if item.start_date <= current_date <= item.end_date
            ]
            if len(active) > self.MAX_DAILY_ACTIVITIES:
                raise WorkPlanError(
                    f"El {current_date:%d/%m/%Y} tiene mas de tres actividades: "
                    + ", ".join(active)
                )
            current_date += timedelta(days=1)
        phase_limits = []
        for phase in self.PHASE_CODES:
            items = [item for item in schedule if item.phase == phase]
            phase_limits.append((min(item.start_date for item in items), max(item.end_date for item in items)))
        for previous, current in zip(phase_limits, phase_limits[1:]):
            if current[0] != previous[1] + timedelta(days=1):
                raise WorkPlanError("Las fases deben ser cronologicas y consecutivas.")

    def _complete_workbook_with_excel(
        self,
        path: Path,
        project: dict[str, Any],
        schedule: list[ScheduledActivity],
        talent: dict[str, Any],
        expert: dict[str, Any],
    ) -> None:
        try:
            import pythoncom
            import win32com.client
        except ImportError as error:
            raise WorkPlanError(
                "La generacion local requiere Microsoft Excel y pywin32."
            ) from error

        excel = None
        workbook = None
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=False)
            sheet = workbook.Worksheets(self.SHEET_NAME)

            classification = str(sheet.Range("A8").Value or "Publica").strip()
            sheet.Range("A8").Value = f"{classification}  X"
            sheet.Range("C11").Value = project.get("name") or "N.A."
            project_name_range = sheet.Range("C11:H12")
            project_name_range.WrapText = True
            project_name_range.ShrinkToFit = False
            project_name_range.HorizontalAlignment = -4131
            project_name_range.VerticalAlignment = -4108
            project_name_range.Font.Size = 9
            sheet.Range("R11").Value = project.get("code") or "N.A."
            sheet.Range("R12").Value = f"TRL {project.get('target_trl')}"
            assigned_name = self.short_name(talent.get("name"))
            expert_name = self.short_name(expert.get("name"))

            for item in schedule:
                row = item.row
                sheet.Cells(row, 3).Value = item.category
                sheet.Cells(row, 4).Value = assigned_name
                sheet.Cells(row, 5).Value = expert_name
                sheet.Cells(row, 6).Value = datetime.combine(item.start_date, datetime.min.time())
                sheet.Cells(row, 6).NumberFormat = "dd/mm/yyyy"
                sheet.Cells(row, 7).Value = item.duration
                sheet.Range(sheet.Cells(row, 9), sheet.Cells(row, 39)).ClearContents()
                for day in self._day_numbers(item.start_date, item.end_date):
                    sheet.Cells(row, 8 + day).Value = "X"
                matrix_range = sheet.Range(sheet.Cells(row, 3), sheet.Cells(row, 39))
                matrix_range.HorizontalAlignment = -4108
                matrix_range.VerticalAlignment = -4108
                matrix_range.WrapText = True

            logo_height = 43.5
            logo_width = logo_height * 500 / 485
            target = sheet.Range("A1:AI2")
            left = target.Left + (target.Width - logo_width) / 2
            top = target.Top + (target.Height - logo_height) / 2
            sheet.Shapes.AddPicture(
                str(SENA_LOGO_PATH.resolve()),
                False,
                True,
                left,
                top,
                logo_width,
                logo_height,
            )
            sheet.PageSetup.Orientation = 2
            sheet.PageSetup.PaperSize = 8
            sheet.PageSetup.Zoom = False
            sheet.PageSetup.FitToPagesWide = 1
            sheet.PageSetup.FitToPagesTall = False
            sheet.PageSetup.PrintArea = "$A$1:$AM$48"
            workbook.Worksheets("Listas desplegables").Visible = 0
            workbook.Save()
        except Exception as error:
            raise WorkPlanError(
                f"No fue posible diligenciar la plantilla con Microsoft Excel: {error}"
            ) from error
        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    pass
            pythoncom.CoUninitialize()
    def _complete_workbook(
        self,
        path: Path,
        project: dict[str, Any],
        schedule: list[ScheduledActivity],
        talent: dict[str, Any],
        expert: dict[str, Any],
    ) -> None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(path)
        sheet = workbook[self.SHEET_NAME]
        add_centered_image(sheet, SENA_LOGO_PATH, "A1:AI2", height=58)

        sheet["A8"] = f"{str(sheet['A8'].value or 'Publica').strip()}  X"
        sheet["C11"] = project.get("name") or "N.A."
        sheet["C11"].alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=False
        )
        project_name_font = copy(sheet["C11"].font)
        project_name_font.sz = 9
        sheet["C11"].font = project_name_font
        sheet["R11"] = project.get("code") or "N.A."
        sheet["R12"] = f"TRL {project.get('target_trl')}"
        assigned_name = self.short_name(talent.get("name"))
        expert_name = self.short_name(expert.get("name"))

        for item in schedule:
            row = item.row
            sheet.cell(row, 3).value = item.category
            sheet.cell(row, 4).value = assigned_name
            sheet.cell(row, 5).value = expert_name
            date_cell = sheet.cell(row, 6)
            date_cell.value = item.start_date
            date_cell.number_format = "dd/mm/yyyy"
            sheet.cell(row, 7).value = item.duration
            for column in range(9, 40):
                sheet.cell(row, column).value = None
            expected_days = self._day_numbers(item.start_date, item.end_date)
            for day in expected_days:
                sheet.cell(row, 8 + day).value = "X"
            for column in range(3, 40):
                cell = sheet.cell(row, column)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=True,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent,
                )

        self._validate_marks(sheet, schedule)
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
        sheet.print_area = "A1:AM48"
        workbook["Listas desplegables"].sheet_state = "hidden"
        workbook.active = workbook.sheetnames.index(self.SHEET_NAME)
        workbook.save(path)

    @staticmethod
    @lru_cache(maxsize=1)
    def _load_descriptions() -> dict[str, str]:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(WORK_PLAN_TEMPLATE, read_only=False, data_only=False)
        sheet = workbook[WorkPlanService.SHEET_NAME]
        descriptions = {
            code: str(sheet.cell(row, 2).value or "").strip()
            for code, row in WorkPlanService.ACTIVITY_ROWS.items()
        }
        workbook.close()
        if any(not value for value in descriptions.values()):
            raise WorkPlanError("La plantilla contiene actividades sin descripcion.")
        return descriptions

    def _schedule_execution_middle(
        self,
        codes: tuple[str, ...],
        descriptions: dict[str, str],
        phase: int,
        phase_start: date,
        phase_end: date,
    ) -> list[ScheduledActivity]:
        """Distribuye secuencialmente las actividades dentro del bloque de ejecución."""

        interior_start = phase_start + timedelta(days=1)
        interior_days = max(0, (phase_end - phase_start).days - 1)
        if interior_days < len(codes):
            return self._schedule_middle(
                codes, descriptions, phase, phase_start, phase_end
            )
        lengths = self._equal_lengths(interior_days, len(codes))
        result: list[ScheduledActivity] = []
        cursor = interior_start
        for code, length in zip(codes, lengths):
            end = cursor + timedelta(days=length - 1)
            result.append(
                self._activity(code, descriptions, phase, "Desarrollo", cursor, end)
            )
            cursor = end + timedelta(days=1)
        return result
    def _schedule_middle(
        self,
        codes: tuple[str, ...],
        descriptions: dict[str, str],
        phase: int,
        phase_start: date,
        phase_end: date,
    ) -> list[ScheduledActivity]:
        if not codes:
            return []
        day_count = (phase_end - phase_start).days + 1
        capacity = [self.MAX_DAILY_ACTIVITIES] * day_count
        capacity[0] -= 1  # actividad de inicio de fase
        capacity[-1] -= 2  # pareja de cierre de fase
        slots = [index for index, available in enumerate(capacity) for _ in range(available)]
        if len(slots) < len(codes):
            raise WorkPlanError(
                f"La fase {phase} no tiene capacidad para limitar el cronograma a tres actividades por dia."
            )

        if len(codes) == 1:
            selected = [slots[len(slots) // 2]]
        else:
            selected = [
                slots[round(index * (len(slots) - 1) / (len(codes) - 1))]
                for index in range(len(codes))
            ]
        occupancy = [0] * day_count
        occupancy[0] = 1
        occupancy[-1] += 2
        intervals = [[day_index, day_index] for day_index in selected]
        for day_index in selected:
            occupancy[day_index] += 1

        center = (len(intervals) - 1) / 2
        expansion_order = sorted(
            range(len(intervals)), key=lambda index: (abs(index - center), index)
        )
        changed = True
        while changed:
            changed = False
            for index in expansion_order:
                left, right = intervals[index]
                if right - left + 1 >= self.MAX_ACTIVITY_DURATION:
                    continue
                candidates = [
                    day_index
                    for day_index in (right + 1, left - 1)
                    if 0 <= day_index < day_count
                    and occupancy[day_index] < self.MAX_DAILY_ACTIVITIES
                ]
                if candidates:
                    candidate = min(candidates, key=lambda day_index: occupancy[day_index])
                    intervals[index][0] = min(left, candidate)
                    intervals[index][1] = max(right, candidate)
                    occupancy[candidate] += 1
                    changed = True

        return [
            self._activity(
                code,
                descriptions,
                phase,
                "Desarrollo",
                phase_start + timedelta(days=interval[0]),
                phase_start + timedelta(days=interval[1]),
            )
            for code, interval in zip(codes, intervals)
        ]
    def _activity(
        self,
        code: str,
        descriptions: dict[str, str],
        phase: int,
        category: str,
        start_date: date,
        end_date: date,
    ) -> ScheduledActivity:
        return ScheduledActivity(
            code=code,
            description=descriptions[code],
            phase=phase,
            row=self.ACTIVITY_ROWS[code],
            category=category,
            start_date=start_date,
            end_date=end_date,
        )

    def _allocate_phase_lengths(self, total_days: int) -> list[int]:
        raw = [total_days * weight for weight in self.PHASE_WEIGHTS]
        execution_days = round(raw[2])
        execution_days = max(self.MIN_PHASE_LENGTHS[2], execution_days)
        execution_days = min(
            execution_days,
            total_days - sum(self.MIN_PHASE_LENGTHS[index] for index in (0, 1, 3)),
        )
        lengths = [
            max(self.MIN_PHASE_LENGTHS[index], math.floor(raw[index]))
            for index in range(4)
        ]
        lengths[2] = execution_days
        adjustable = (0, 1, 3)
        while sum(lengths) < total_days:
            index = max(adjustable, key=lambda item: raw[item] - lengths[item])
            lengths[index] += 1
        while sum(lengths) > total_days:
            candidates = [
                index
                for index in adjustable
                if lengths[index] > self.MIN_PHASE_LENGTHS[index]
            ]
            if not candidates:
                raise WorkPlanError(
                    "No es posible distribuir las cuatro fases en el periodo indicado."
                )
            index = max(candidates, key=lambda item: lengths[item] - raw[item])
            lengths[index] -= 1
        return lengths

    @staticmethod
    def _equal_lengths(total: int, count: int) -> list[int]:
        quotient, remainder = divmod(total, count)
        return [quotient + (1 if index < remainder else 0) for index in range(count)]
    @staticmethod
    def _day_numbers(start_date: date, end_date: date) -> set[int]:
        return {
            (start_date + timedelta(days=offset)).day
            for offset in range((end_date - start_date).days + 1)
        }

    def _validate_marks(self, sheet: Any, schedule: list[ScheduledActivity]) -> None:
        for item in schedule:
            expected = self._day_numbers(item.start_date, item.end_date)
            actual = {
                column - 8
                for column in range(9, 40)
                if str(sheet.cell(item.row, column).value or "").strip().upper() == "X"
            }
            if actual != expected:
                raise WorkPlanError(f"Las marcas X de la actividad {item.code} no corresponden a sus fechas.")

    @staticmethod
    def short_name(value: Any) -> str:
        parts = str(value or "").strip().split()
        if len(parts) < 2:
            return parts[0] if parts else "N.A."
        surname_index = -2 if len(parts) >= 3 else -1
        return f"{parts[0]} {parts[surname_index]}"

    @staticmethod
    def _as_date(value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str) and value.strip():
            try:
                return date.fromisoformat(value.strip())
            except ValueError:
                return None
        return None

    @staticmethod
    def _safe_filename(value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
        return cleaned.strip("_") or "proyecto"
