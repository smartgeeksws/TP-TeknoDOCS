"""Generacion temporal del formato GCDTP-F-019 en PDF."""

from __future__ import annotations

import re
import shutil
import tempfile
import warnings
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

from config.settings import TECHNOLOGY_BASE_PROJECT_TEMPLATE
from services.document_generation.spreadsheet_pdf import (
    SpreadsheetPdfError,
    convert_xlsx_to_pdf,
)


class TechnologyBaseProjectError(RuntimeError):
    """Error controlado al producir el formato GCDTP-F-019."""


class TechnologyBaseProjectService:
    """Completa una copia temporal de la plantilla y retorna solo el PDF."""

    TECHNO_PARK = "Tecnoparque Nodo Angostura"
    REGIONAL_CENTER = (
        "Regional Huila - Centro de Formaci\u00f3n Agroindustrial La Angostura"
    )
    RESPONSIBLE_ROLE = (
        "Experto de Innovaci\u00f3n y Desarrollo Tecnol\u00f3gico de Tecnoparque"
    )

    def generate(
        self,
        project: dict[str, Any],
        content: dict[str, str],
    ) -> tuple[bytes, str]:
        self._validate(project, content)
        with tempfile.TemporaryDirectory(prefix="tp_pbt_") as temp_directory:
            directory = Path(temp_directory)
            xlsx_path = directory / "proyecto_base_tecnologica.xlsx"
            pdf_path = directory / "proyecto_base_tecnologica.pdf"
            shutil.copy2(TECHNOLOGY_BASE_PROJECT_TEMPLATE, xlsx_path)
            self._complete_workbook(xlsx_path, project, content)
            try:
                convert_xlsx_to_pdf(xlsx_path, pdf_path)
            except SpreadsheetPdfError as error:
                raise TechnologyBaseProjectError(str(error)) from error
            data = pdf_path.read_bytes()
        safe_code = self._safe_filename(project.get("code") or "proyecto")
        return data, f"Proyecto Base Tecnol\u00f3gica {safe_code}.pdf"

    def _complete_workbook(
        self,
        path: Path,
        project: dict[str, Any],
        content: dict[str, str],
    ) -> None:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(path)
        sheet = workbook["GCDTP-F-019"]
        self._add_fourth_objective(sheet)

        expert = project["expert"]
        fixed_values = {
            "B11": self.TECHNO_PARK,
            "B12": self.REGIONAL_CENTER,
            "B13": expert["name"],
            "E13": self.RESPONSIBLE_ROLE,
            "B15": project["code"],
            "B17": project["name"],
            "B18": project["technology_line"],
            "B19": project["initial_trl"],
            "E19": project["target_trl"],
            "B20": content["main_approach"],
            "B22": content["general_objective"],
            "B23": content["specific_objective_1"],
            "B24": content["specific_objective_2"],
            "B25": content["specific_objective_3"],
            "B26": content["specific_objective_4"],
            "B27": content["scope"],
        }
        for coordinate, value in fixed_values.items():
            self._set_wrapped(sheet, coordinate, value)

        sheet["A8"] = "P\u00fablica  X"
        self._fill_people(sheet, 29, project.get("talents", []))
        self._fill_intellectual_owners(sheet, 34, 39, project)
        self._fill_research_groups(sheet, 44, project)
        for row in (20, 22, 23, 24, 25, 26, 27):
            self._fit_text_row(sheet, row)

        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_area = "A1:F49"
        sheet.sheet_properties.pageSetUpPr.autoPageBreaks = False
        workbook["Listas desplegables"].sheet_state = "hidden"
        workbook.active = workbook.sheetnames.index("GCDTP-F-019")
        workbook.save(path)

    @staticmethod
    def _add_fourth_objective(sheet: Any) -> None:
        downstream_merges = [
            merged
            for merged in list(sheet.merged_cells.ranges)
            if merged.min_row >= 27
        ]
        for merged in downstream_merges:
            sheet.unmerge_cells(str(merged))
        sheet.unmerge_cells("B26:F26")

        scope_cells = []
        objective_cells = []
        for column in range(1, 7):
            scope_cells.append(TechnologyBaseProjectService._cell_snapshot(sheet.cell(26, column)))
            objective_cells.append(TechnologyBaseProjectService._cell_snapshot(sheet.cell(25, column)))
        scope_height = sheet.row_dimensions[26].height
        objective_height = sheet.row_dimensions[25].height
        row_dimensions = {
            row: (sheet.row_dimensions[row].height, sheet.row_dimensions[row].hidden)
            for row in range(27, sheet.max_row + 1)
        }

        sheet.insert_rows(27)
        for merged in downstream_merges:
            sheet.merge_cells(
                start_row=merged.min_row + 1,
                start_column=merged.min_col,
                end_row=merged.max_row + 1,
                end_column=merged.max_col,
            )
        for row in range(sheet.max_row, 27, -1):
            old = row_dimensions.get(row - 1)
            if old:
                sheet.row_dimensions[row].height = old[0]
                sheet.row_dimensions[row].hidden = old[1]

        for column, snapshot in enumerate(scope_cells, start=1):
            TechnologyBaseProjectService._restore_cell(sheet.cell(27, column), snapshot)
        sheet.merge_cells("B27:F27")
        sheet.row_dimensions[27].height = scope_height

        for column, snapshot in enumerate(objective_cells, start=1):
            TechnologyBaseProjectService._restore_cell(sheet.cell(26, column), snapshot)
        sheet["A26"] = "Objetivo Espec\u00edfico 4"
        sheet["B26"] = None
        sheet.merge_cells("B26:F26")
        sheet.row_dimensions[26].height = objective_height

    @staticmethod
    def _cell_snapshot(cell: Any) -> dict[str, Any]:
        return {
            "value": None if isinstance(cell, MergedCell) else cell.value,
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
            "alignment": copy(cell.alignment),
        }

    @staticmethod
    def _restore_cell(cell: Any, snapshot: dict[str, Any]) -> None:
        cell.value = snapshot["value"]
        cell._style = copy(snapshot["style"])
        cell.number_format = snapshot["number_format"]
        cell.protection = copy(snapshot["protection"])
        cell.alignment = copy(snapshot["alignment"])

    @staticmethod
    def _fill_people(sheet: Any, start_row: int, people: list[dict[str, Any]]) -> None:
        values = people or [{"name": "N.A.", "document_number": "N.A."}]
        for offset in range(4):
            row = start_row + offset
            visible = offset < len(values)
            sheet.row_dimensions[row].hidden = not visible
            if visible:
                person = values[offset]
                sheet.cell(row, 2).value = person.get("name") or "N.A."
                document = person.get("document_number") or "N.A."
                document_type = person.get("document_type")
                sheet.cell(row, 5).value = (
                    f"{document_type} {document}" if document_type else document
                )

    def _fill_intellectual_owners(
        self,
        sheet: Any,
        talent_start: int,
        company_start: int,
        project: dict[str, Any],
    ) -> None:
        company = project.get("company")
        titular = next(
            (item for item in project.get("talents", []) if item.get("role") == "titular"),
            None,
        )
        if company:
            talent_values = [{"name": "N.A.", "document_number": "N.A."}]
            company_values = [(company.get("legal_name") or "N.A.", company.get("nit") or "N.A.")]
        else:
            talent_values = [titular] if titular else [{"name": "N.A.", "document_number": "N.A."}]
            company_values = [("N.A.", "N.A.")]
        self._fill_people(sheet, talent_start, talent_values)
        self._fill_pairs(sheet, company_start, company_values)

    @staticmethod
    def _fill_research_groups(sheet: Any, start_row: int, project: dict[str, Any]) -> None:
        groups = project.get("research_groups") or []
        if not groups and project.get("research_group_name"):
            groups = [{"name": project["research_group_name"], "minciencias_code": None}]
        values = [
            (group.get("name") or "N.A.", group.get("minciencias_code") or "N.A.")
            for group in groups
        ] or [("N.A.", "N.A.")]
        TechnologyBaseProjectService._fill_pairs(sheet, start_row, values)

    @staticmethod
    def _fill_pairs(sheet: Any, start_row: int, values: list[tuple[str, str]]) -> None:
        for offset in range(4):
            row = start_row + offset
            visible = offset < len(values)
            sheet.row_dimensions[row].hidden = not visible
            if visible:
                sheet.cell(row, 2).value, sheet.cell(row, 5).value = values[offset]

    @staticmethod
    def _set_wrapped(sheet: Any, coordinate: str, value: Any) -> None:
        cell = sheet[coordinate]
        cell.value = "" if value is None else str(value)
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical="top",
            text_rotation=cell.alignment.text_rotation,
            wrap_text=True,
            shrink_to_fit=False,
            indent=cell.alignment.indent,
        )

    @staticmethod
    def _fit_text_row(sheet: Any, row: int) -> None:
        text = str(sheet.cell(row, 2).value or "")
        estimated_lines = max(1, sum(max(1, (len(line) + 84) // 85) for line in text.splitlines()))
        sheet.row_dimensions[row].height = max(
            sheet.row_dimensions[row].height or 15,
            15 * (estimated_lines + 1),
        )

    @staticmethod
    def _validate(project: dict[str, Any], content: dict[str, str]) -> None:
        required_project = {
            "code": "codigo",
            "name": "nombre",
            "technology_line": "linea tecnologica",
            "initial_trl": "TRL inicial",
            "target_trl": "TRL objetivo",
            "expert": "experto Tecnoparque",
        }
        missing = [label for field, label in required_project.items() if not project.get(field)]
        if missing:
            raise TechnologyBaseProjectError(
                "Faltan datos del proyecto: " + ", ".join(missing) + "."
            )
        missing_content = [key for key, value in content.items() if not str(value).strip()]
        if missing_content:
            raise TechnologyBaseProjectError("Completa todos los textos antes de generar el PDF.")
        if not TECHNOLOGY_BASE_PROJECT_TEMPLATE.is_file():
            raise TechnologyBaseProjectError("No se encontro la plantilla GCDTP-F-019.")

    @staticmethod
    def _safe_filename(value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
        return cleaned.strip("_") or "proyecto"
